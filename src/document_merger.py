from flask import Blueprint, render_template, request, send_file
import json
import os
import mimetypes
from io import BytesIO
from openpyxl import load_workbook
from docx import Document

document_merger = Blueprint('document_merger', __name__)

def replace_text_in_excel(file_path, replacements):
    workbook = load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.value = replace_text(cell.value, replacements)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def replace_text_in_docx(file_path, replacements):
    document = Document(file_path)
    for paragraph in document.paragraphs:
        for run in paragraph.runs:
            run.text = replace_text(run.text, replacements)
    output = BytesIO()
    document.save(output)
    output.seek(0)
    return output

def replace_text(document_text, replacements):
    for replacement in replacements:
        key = replacement.get('key', '')
        value = replacement.get('value', '')
        document_text = document_text.replace(key, value)
    return document_text

@document_merger.route('/')
def index():
    return render_template('document_merger.html')

@document_merger.route('/', methods=['POST'])
def merge_documents():
    document_file = request.files['file']
    replacements_raw = request.form.get('replacements', '')

    try:
        replacements = json.loads(replacements_raw)
        if not isinstance(replacements, list):
            raise ValueError("Invalid replacements format")
    except Exception as e:
        return f"Error: {str(e)}"

    # Determine the MIME type based on the file extension
    mime_type, _ = mimetypes.guess_type(document_file.filename)
    if mime_type is None:
        mime_type = 'application/octet-stream'  # Default to binary if MIME type is unknown

    # Perform text replacements based on the file extension
    if document_file.filename.lower().endswith('.xlsx'):
        modified_content = replace_text_in_excel(document_file, replacements)
    elif document_file.filename.lower().endswith('.docx'):
        modified_content = replace_text_in_docx(document_file, replacements)
    else:
        return "Unsupported file format"

    # Return the modified document with the same file name and extension as the uploaded document
    return send_file(
        modified_content,
        as_attachment=True,
        download_name=document_file.filename,
        mimetype=mime_type
    )
